#!/usr/bin/env python3
"""
Create the contact form Excel file with headers.
This script creates yosoc_messages.xlsx with the proper structure.
"""

from openpyxl import Workbook
import os

# Configuration
CONTACT_EXCEL_FILE_PATH = 'yosoc_messages.xlsx'
CONTACT_SHEET_NAME = 'Messages'
CONTACT_HEADERS = ['Timestamp', 'Name', 'Email', 'Subject', 'Message']

def create_contact_excel_file():
    """Create contact Excel file with headers"""
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = CONTACT_SHEET_NAME
        
        # Add headers
        for col, header in enumerate(CONTACT_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Set column widths for better readability
        column_widths = [20, 25, 30, 40, 50]  # Adjust as needed
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        
        # Save the workbook
        wb.save(CONTACT_EXCEL_FILE_PATH)
        print(f"✅ Created contact Excel file: {CONTACT_EXCEL_FILE_PATH}")
        print(f"📁 Location: {os.path.abspath(CONTACT_EXCEL_FILE_PATH)}")
        print(f"📊 Sheet name: {CONTACT_SHEET_NAME}")
        print(f"📋 Headers: {', '.join(CONTACT_HEADERS)}")
        return True
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {str(e)}")
        return False

if __name__ == '__main__':
    print("Creating contact form Excel file...")
    if create_contact_excel_file():
        print("\n🎉 Excel file created successfully!")
        print("You can now open it in Excel or any spreadsheet application.")
    else:
        print("\n💥 Failed to create Excel file.")
