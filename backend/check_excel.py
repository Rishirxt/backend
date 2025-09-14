#!/usr/bin/env python3
"""
Check the contents of the yosoc_applications.xlsx file
"""

from openpyxl import load_workbook

def check_excel_file():
    """Check the Excel file contents"""
    try:
        wb = load_workbook('yosoc_applications.xlsx')
        ws = wb['Applications']
        
        print("Y-SoC Applications Excel File Contents")
        print("="*50)
        
        print("Headers:")
        headers = [ws.cell(1, col).value for col in range(1, 14)]
        print(headers)
        
        print(f"\nTotal rows: {ws.max_row}")
        print(f"Data rows: {ws.max_row - 1}")
        
        if ws.max_row > 1:
            print("\nData rows:")
            for row in range(2, ws.max_row + 1):
                row_data = [ws.cell(row, col).value for col in range(1, 14)]
                print(f"Row {row-1}: {row_data}")
        else:
            print("No data rows found")
            
    except Exception as e:
        print(f"Error reading Excel file: {str(e)}")

if __name__ == "__main__":
    check_excel_file()
