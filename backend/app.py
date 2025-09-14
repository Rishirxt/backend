from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
import os
from datetime import datetime
import json
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configuration
EXCEL_FILE_PATH = 'yosoc_applications.xlsx'
SHEET_NAME = 'Applications'

# Contact form configuration
CONTACT_EXCEL_FILE_PATH = 'yosoc_messages.xlsx'
CONTACT_SHEET_NAME = 'Messages'
CONTACT_HEADERS = ['Timestamp', 'Name', 'Email', 'Subject', 'Message']

# Define the required headers for the Excel file (in correct order)
HEADERS = [
    "First Name", "Last Name", "Email", "Country", "Role", 
    "Experience Level", "Technical Skills", "Why Join", 
    "GitHub", "Portfolio", "Time Commitment", 
    "Agree Terms", "Updates Subscription"
]

def create_application_excel_file_if_not_exists():
    """Create application Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        
        # Add headers
        for col, header in enumerate(HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Save the workbook
        wb.save(EXCEL_FILE_PATH)
        print(f"Created new application Excel file: {EXCEL_FILE_PATH}")

def check_email_exists(email):
    """Check if email already exists in the Excel file"""
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return False
        
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb[SHEET_NAME]
        
        # Check all rows (starting from row 2, since row 1 is headers)
        for row in range(2, ws.max_row + 1):
            email_cell = ws.cell(row=row, column=3)  # Email is in column 3
            if email_cell.value and email_cell.value.lower() == email.lower():
                return True
        
        return False
    except Exception as e:
        print(f"Error checking email existence: {str(e)}")
        return False

def append_application_to_excel(data):
    """Append new application data to the Excel file using openpyxl"""
    try:
        # Create file if it doesn't exist
        create_application_excel_file_if_not_exists()
        
        # Check if email already exists
        if check_email_exists(data.get('email', '')):
            return False, "Email already registered"
        
        # Load existing workbook
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb[SHEET_NAME]
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Prepare data for the new row (in correct order)
        row_data = [
            data.get('first_name', ''),
            data.get('last_name', ''),
            data.get('email', ''),
            data.get('country', ''),
            data.get('role', ''),
            data.get('experience_level', ''),
            data.get('technical_skills', ''),
            data.get('why_join', ''),
            data.get('github', ''),
            data.get('portfolio', ''),
            data.get('time_commitment', ''),
            data.get('agree_terms', False),
            data.get('updates_subscription', False)
        ]
        
        # Add data to the worksheet
        for col, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col, value=value)
        
        # Save the workbook
        wb.save(EXCEL_FILE_PATH)
        print(f"Application data saved to Excel: {EXCEL_FILE_PATH}")
        return True, "Application stored successfully"
        
    except Exception as e:
        print(f"Error saving application data to Excel: {str(e)}")
        return False, f"Error saving application: {str(e)}"

def create_contact_excel_file_if_not_exists():
    """Create contact Excel file with headers if it doesn't exist"""
    if not os.path.exists(CONTACT_EXCEL_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = CONTACT_SHEET_NAME
        
        # Add headers
        for col, header in enumerate(CONTACT_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Save the workbook
        wb.save(CONTACT_EXCEL_FILE_PATH)
        print(f"Created new contact Excel file: {CONTACT_EXCEL_FILE_PATH}")

def save_contact_to_excel(data):
    """Save contact form data to Excel file using openpyxl"""
    try:
        # Create file if it doesn't exist
        create_contact_excel_file_if_not_exists()
        
        # Load existing workbook
        wb = load_workbook(CONTACT_EXCEL_FILE_PATH)
        ws = wb[CONTACT_SHEET_NAME]
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Prepare data for the new row
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row_data = [
            timestamp,
            data.get('name', ''),
            data.get('email', ''),
            data.get('subject', ''),
            data.get('message', '')
        ]
        
        # Add data to the worksheet
        for col, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col, value=value)
        
        # Save the workbook
        wb.save(CONTACT_EXCEL_FILE_PATH)
        print(f"Contact form data saved to Excel: {CONTACT_EXCEL_FILE_PATH}")
        return True
        
    except Exception as e:
        print(f"Error saving contact data to Excel: {str(e)}")
        return False

@app.route('/submit', methods=['POST'])
def submit_application():
    """Handle form submission and store data in Excel"""
    try:
        # Get JSON data from request
        data = request.get_json()
        
        if not data:
            return jsonify({
                "status": "error",
                "message": "No data provided"
            }), 400
        
        # Validate required fields
        required_fields = ['first_name', 'last_name', 'email', 'country', 'role', 
                          'experience_level', 'technical_skills', 'why_join', 
                          'time_commitment', 'agree_terms']
        
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                "status": "error",
                "message": f"Missing required fields: {', '.join(missing_fields)}"
            }), 400
        
        # Append data to Excel and check for email uniqueness
        success, message = append_application_to_excel(data)
        
        if success:
            return jsonify({
                "success": True,
                "message": message
            }), 200
        else:
            # Check if it's an email duplicate error
            if "already registered" in message:
                return jsonify({
                    "success": False,
                    "message": "Email already registered"
                }), 400
            else:
                return jsonify({
                    "success": False,
                    "message": message
                }), 500
            
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Server error: {str(e)}"
        }), 500

@app.route('/contact', methods=['POST'])
def submit_contact():
    """Handle contact form submission and store data in Excel"""
    try:
        # Get JSON data from request
        data = request.get_json()
        
        if not data:
            return jsonify({
                "status": "error",
                "message": "No data provided"
            }), 400
        
        # Validate required fields
        required_fields = ['name', 'email', 'subject', 'message']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                "status": "error",
                "message": f"Missing required fields: {', '.join(missing_fields)}"
            }), 400
        
        # Save contact data to Excel
        if save_contact_to_excel(data):
            return jsonify({
                "status": "success",
                "message": "Contact form data saved successfully"
            }), 200
        else:
            return jsonify({
                "status": "error",
                "message": "Failed to save contact form data"
            }), 500
            
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Server error: {str(e)}"
        }), 500

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        "status": "success",
        "message": "Flask backend is running"
    }), 200

@app.route('/applications', methods=['GET'])
def get_applications():
    """Get all applications (for admin purposes)"""
    try:
        if not os.path.exists(EXCEL_FILE_PATH):
            return jsonify({
                "status": "success",
                "data": [],
                "message": "No applications found"
            }), 200
        
        df = pd.read_excel(EXCEL_FILE_PATH, sheet_name=SHEET_NAME)
        applications = df.to_dict('records')
        
        return jsonify({
            "status": "success",
            "data": applications,
            "count": len(applications)
        }), 200
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Error retrieving applications: {str(e)}"
        }), 500

if __name__ == '__main__':
    # Create the Excel file if it doesn't exist when starting the server
    create_application_excel_file_if_not_exists()
    
    print("Starting Flask backend server...")
    print(f"Excel file will be created at: {os.path.abspath(EXCEL_FILE_PATH)}")
    print("Available endpoints:")
    print("  POST /submit - Submit application form")
    print("  POST /contact - Submit contact form")
    print("  GET /health - Health check")
    print("  GET /applications - Get all applications")
    
    # Get port from environment variable (for Render) or default to 5000
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
